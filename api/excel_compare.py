from fastapi import APIRouter, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook, Workbook
import io
import math
import re

router = APIRouter()


def _norm_label(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _is_number(v) -> bool:
    return isinstance(v, (int, float)) and not (isinstance(v, float) and math.isnan(v))


def _to_number_if_possible(v):
    if _is_number(v):
        return float(v)
    if isinstance(v, str):
        s = "".join(v.split())  # remove all whitespace
        if s == "":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _equal(v1, v2) -> bool:
    if v1 in (None, "") and v2 in (None, ""):
        return True
    n1 = _to_number_if_possible(v1)
    n2 = _to_number_if_possible(v2)
    if n1 is not None and n2 is not None:
        return abs(n1 - n2) < 1e-9
    s1 = "".join(str(v1).split())
    s2 = "".join(str(v2).split())
    return s1 == s2


def _extract_table_auto(ws):
    def build(header_row: int, header_col: int):
        row_map: dict[str, int] = {}
        for r in range(header_row + 1, ws.max_row + 1):
            label = _norm_label(ws.cell(row=r, column=header_col).value)
            if label:
                row_map.setdefault(label, r)

        col_map: dict[str, int] = {}
        for c in range(header_col + 1, ws.max_column + 1):
            label = _norm_label(ws.cell(row=header_row, column=c).value)
            if label:
                col_map.setdefault(label, c)

        values: dict[tuple[str, str], object] = {}
        for r_label, r_idx in row_map.items():
            for c_label, c_idx in col_map.items():
                values[(r_label, c_label)] = ws.cell(row=r_idx, column=c_idx).value
        return row_map, col_map, values

    row_map, col_map, values = build(1, 1)
    if not row_map or not col_map:
        max_r = min(ws.max_row, 50)
        max_c = min(ws.max_column, 50)
        best_hr, best_hr_count = 1, -1
        for r in range(1, max_r + 1):
            seen = set()
            cnt = 0
            for c in range(1, max_c + 1):
                v = _norm_label(ws.cell(row=r, column=c).value)
                if v and v not in seen:
                    seen.add(v)
                    cnt += 1
            if cnt > best_hr_count:
                best_hr_count = cnt
                best_hr = r

        best_hc, best_hc_count = 1, -1
        for c in range(1, max_c + 1):
            seen = set()
            cnt = 0
            for r in range(1, max_r + 1):
                v = _norm_label(ws.cell(row=r, column=c).value)
                if v and v not in seen:
                    seen.add(v)
                    cnt += 1
            if cnt > best_hc_count:
                best_hc_count = cnt
                best_hc = c

        row_map, col_map, values = build(best_hr, best_hc)

    return {
        "row_labels": set(row_map.keys()),
        "col_labels": set(col_map.keys()),
        "values": values,
    }


@router.post("/excel/compare/preview")
async def excel_compare_preview(excel: UploadFile = File(...)):
    data = await excel.read()
    wb = load_workbook(io.BytesIO(data), data_only=True)

    names = wb.sheetnames
    if len(names) < 2:
        return [{"message": "No other sheet to compare"}]

    base_name = names[0]
    base = _extract_table_auto(wb[base_name])
    results = []
    for other_name in names[1:]:
        other = _extract_table_auto(wb[other_name])
        # header diffs
        for r in sorted(list(base["row_labels"] - other["row_labels"])):
            results.append({
                "sheet_left": base_name,
                "sheet_right": other_name,
                "diff_type": "missing_row_in_right",
                "row_label": r,
                "col_label": "",
                "left": "<row exists>",
                "right": "<missing>",
            })
        for r in sorted(list(other["row_labels"] - base["row_labels"])):
            results.append({
                "sheet_left": base_name,
                "sheet_right": other_name,
                "diff_type": "missing_row_in_left",
                "row_label": r,
                "col_label": "",
                "left": "<missing>",
                "right": "<row exists>",
            })
        for c in sorted(list(base["col_labels"] - other["col_labels"])):
            results.append({
                "sheet_left": base_name,
                "sheet_right": other_name,
                "diff_type": "missing_col_in_right",
                "row_label": "",
                "col_label": c,
                "left": "<col exists>",
                "right": "<missing>",
            })
        for c in sorted(list(other["col_labels"] - base["col_labels"])):
            results.append({
                "sheet_left": base_name,
                "sheet_right": other_name,
                "diff_type": "missing_col_in_left",
                "row_label": "",
                "col_label": c,
                "left": "<missing>",
                "right": "<col exists>",
            })

        # cell diffs
        common_rows = base["row_labels"] & other["row_labels"]
        common_cols = base["col_labels"] & other["col_labels"]
        for r in sorted(common_rows):
            for c in sorted(common_cols):
                vl = base["values"].get((r, c))
                vr = other["values"].get((r, c))
                if not _equal(vl, vr):
                    results.append({
                        "sheet_left": base_name,
                        "sheet_right": other_name,
                        "diff_type": "cell",
                        "row_label": r,
                        "col_label": c,
                        "left": "" if vl is None else str(vl),
                        "right": "" if vr is None else str(vr),
                    })

    if not results:
        return [{"message": "No differences"}]
    return results


@router.post("/excel/compare/excel")
async def excel_compare_excel(excel: UploadFile = File(...)):
    data = await excel.read()
    wb = load_workbook(io.BytesIO(data), data_only=True)

    out_wb = Workbook()
    ws = out_wb.active
    ws.title = "mismatches"
    ws.append(["sheet_left", "sheet_right", "diff_type", "row_label", "col_label", "left", "right"])

    names = wb.sheetnames
    if len(names) < 2:
        ws.append(["", "", "info", "", "", "No other sheet to compare", ""])
        stream = io.BytesIO()
        out_wb.save(stream)
        stream.seek(0)
        return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=excel-compare.xlsx"})

    base_name = names[0]
    base = _extract_table_auto(wb[base_name])
    total = 0
    for other_name in names[1:]:
        other = _extract_table_auto(wb[other_name])

        for r in sorted(list(base["row_labels"] - other["row_labels"])):
            ws.append([base_name, other_name, "missing_row_in_right", r, "", "<row exists>", "<missing>"])
            total += 1
        for r in sorted(list(other["row_labels"] - base["row_labels"])):
            ws.append([base_name, other_name, "missing_row_in_left", r, "", "<missing>", "<row exists>"])
            total += 1
        for c in sorted(list(base["col_labels"] - other["col_labels"])):
            ws.append([base_name, other_name, "missing_col_in_right", "", c, "<col exists>", "<missing>"])
            total += 1
        for c in sorted(list(other["col_labels"] - base["col_labels"])):
            ws.append([base_name, other_name, "missing_col_in_left", "", c, "<missing>", "<col exists>"])
            total += 1

        common_rows = base["row_labels"] & other["row_labels"]
        common_cols = base["col_labels"] & other["col_labels"]
        for r in sorted(common_rows):
            for c in sorted(common_cols):
                vl = base["values"].get((r, c))
                vr = other["values"].get((r, c))
                if not _equal(vl, vr):
                    ws.append([base_name, other_name, "cell", r, c, "" if vl is None else str(vl), "" if vr is None else str(vr)])
                    total += 1

    if total == 0:
        ws.append(["", "", "info", "", "", "No differences", ""])  # No mismatches

    # Append original sheets (values only)
    invalid_chars = re.compile(r"[\\/*?:\[\]]")

    def _safe(n: str) -> str:
        n2 = invalid_chars.sub("", n)
        return n2[:31] or "Sheet"

    def _unique(base_name: str, existing: set[str]) -> str:
        name = _safe(base_name)
        if name not in existing:
            return name
        i = 1
        while True:
            suffix = f"_{i}"
            cand = (name[: 31 - len(suffix)] + suffix) if len(name) + len(suffix) > 31 else name + suffix
            if cand not in existing:
                return cand
            i += 1

    exist = {s.title for s in out_wb.worksheets}
    for orig in wb.sheetnames:
        src = wb[orig]
        dst_title = _unique(f"orig_{orig}", exist)
        exist.add(dst_title)
        dst = out_wb.create_sheet(title=dst_title)
        for r in range(1, src.max_row + 1):
            for c in range(1, src.max_column + 1):
                dst.cell(row=r, column=c, value=src.cell(row=r, column=c).value)

    stream = io.BytesIO()
    out_wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=excel-compare.xlsx"})

