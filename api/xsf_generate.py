from fastapi import APIRouter, UploadFile, File, Form
from fastapi.responses import StreamingResponse
import io
from openpyxl import load_workbook
from urllib.parse import quote

router = APIRouter()


@router.post("/xsf/generate")
async def generate_xsf(
    excel: UploadFile = File(...),
    sheet: str | None = Form(None),
    filename: str | None = Form(None),
):
    data = await excel.read()
    stream = io.BytesIO(data)
    wb = load_workbook(stream, data_only=True)

    # Choose sheet: specific by name if provided, else first sheet
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        ws = wb[wb.sheetnames[0]]

    # Build XSF content: each row as one line, cells concatenated as-is
    lines: list[str] = []
    for row in ws.iter_rows(values_only=True):
        line = ''.join([str(cell) if cell is not None else '' for cell in row])
        lines.append(line)

    content = ("\n".join(lines)).encode("utf-8")

    out_name = (filename or excel.filename or "output").rsplit('.', 1)[0] + ".xsf"

    # Build Content-Disposition with ASCII fallback and RFC 5987 filename*
    def _ascii_fallback(name: str) -> str:
        base = []
        for ch in name:
            o = ord(ch)
            if (48 <= o <= 57) or (65 <= o <= 90) or (97 <= o <= 122) or ch in ("-", "_", "."):
                base.append(ch)
            else:
                if o < 128:
                    base.append("_")
                else:
                    base.append("_")
        fb = "".join(base) or "output.xsf"
        # ensure extension
        if not fb.lower().endswith(".xsf"):
            fb += ".xsf"
        return fb

    ascii_name = _ascii_fallback(out_name)
    cd_value = f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(out_name)}"

    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/octet-stream",
        headers={
            "Content-Disposition": cd_value
        },
    )
