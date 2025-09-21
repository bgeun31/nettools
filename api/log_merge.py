from fastapi import APIRouter, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from typing import List
import io
import os
import re
from openpyxl import Workbook

router = APIRouter()


invalid_chars = re.compile(r"[\\/*?:\[\]]")


def natural_sort_key(text: str):
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", text)]


@router.post("/merge/preview")
async def merge_preview(files: List[UploadFile] = File(...)):
    # Sort files by natural order of filename
    file_entries = sorted(files, key=lambda f: natural_sort_key(f.filename or ""))
    results = []
    for f in file_entries:
        content = (await f.read()).decode("utf-8", errors="ignore")
        raw_name = os.path.splitext(os.path.basename(f.filename or ""))[0]
        safe_name = invalid_chars.sub("", raw_name)[:31] or "Sheet"
        results.append({
            "filename": f.filename,
            "sheet": safe_name,
            "lines": len(content.splitlines()),
        })
    return results


@router.post("/merge/excel")
async def merge_excel(
    files: List[UploadFile] = File(...),
    output_name: str = Form("merged_logs.xlsx"),
):
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    # Sort files by natural order of filename
    file_entries = sorted(files, key=lambda f: natural_sort_key(f.filename or ""))

    existing_names = set()
    for f in file_entries:
        content = (await f.read()).decode("utf-8", errors="ignore")
        lines = content.splitlines()
        raw_name = os.path.splitext(os.path.basename(f.filename or ""))[0]
        safe_name = invalid_chars.sub("", raw_name)[:31] or "Sheet"
        # ensure unique within 31 chars
        original = safe_name
        count = 1
        while safe_name in existing_names:
            suffix = f"_{count}"
            safe_name = (original[: max(0, 31 - len(suffix))] + suffix) or f"Sheet_{count}"
            count += 1
        existing_names.add(safe_name)

        ws = wb.create_sheet(title=safe_name)
        for idx, line in enumerate(lines, 1):
            ws.cell(row=idx, column=1, value=line)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={output_name or 'merged_logs.xlsx'}"
        },
    )

