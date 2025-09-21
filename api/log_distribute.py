from fastapi import APIRouter, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from typing import List
import io
import os
import re
import zipfile
from openpyxl import load_workbook

router = APIRouter()

invalid_chars = re.compile(r"[\\/*?:\[\]]")


def _safe_sheet_name(name: str) -> str:
    safe = invalid_chars.sub("", name)[:31]
    return safe or "Sheet"


@router.post("/distribute/preview")
async def distribute_preview(excel: UploadFile = File(...)):
    data = await excel.read()
    stream = io.BytesIO(data)
    wb = load_workbook(stream, data_only=True)
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines = 0
        for row in ws.iter_rows(values_only=True):
            line = ''.join([str(cell) if cell is not None else '' for cell in row])
            # count even if empty line to reflect file length semantics
            lines += 1
        results.append({
            "sheet": _safe_sheet_name(sheet_name),
            "original_sheet": sheet_name,
            "lines": lines,
        })
    return results


@router.post("/distribute/zip")
async def distribute_zip(
    excel: UploadFile = File(...),
    format: str = Form("txt"),
):
    fmt = (format or "txt").lower()
    if fmt not in {"txt", "log"}:
        fmt = "txt"

    data = await excel.read()
    stream = io.BytesIO(data)
    wb = load_workbook(stream, data_only=True)

    zip_stream = io.BytesIO()
    with zipfile.ZipFile(zip_stream, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            safe_name = _safe_sheet_name(sheet_name)
            txt_filename = f"{safe_name}.{fmt}"
            # Build content
            lines_out = []
            for row in ws.iter_rows(values_only=True):
                line = ''.join([str(cell) if cell is not None else '' for cell in row])
                lines_out.append(line)
            content = ("\n".join(lines_out)).encode("utf-8")
            zf.writestr(txt_filename, content)

    zip_stream.seek(0)
    return StreamingResponse(
        zip_stream,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=distributed-logs.zip"},
    )

